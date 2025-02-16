from openpyxl import load_workbook

wb1 = load_workbook("tabela_przetawna.xlsx")
wb2 = load_workbook("vgsales_formated.xlsx")

ws1 = wb1.active
ws2 = wb2.active

if ws1['A1'].value == ws2['A1'].value:
    print("Wartości A1 są identyczne")
else:
    print(f"Róznica: plik1={ws1['A1'].value}, plik2={ws2['A1'].value}")
