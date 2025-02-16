import glob  # Return a list of paths matching a pathname pattern.
from openpyxl import load_workbook

folder_path = "./"

file_list = glob.glob(folder_path + "*.xlsx")
# file_list = glob.glob("*.xlsx")

print(file_list)
# ['.\\dane_zaktualizowane.xlsx', '.\\openpyxl_optimized.xlsx',
# '.\\tabela_przetawna.xlsx', '.\\vgsales_formated.xlsx', '.\\videogamesales.xlsx', '.\\xlswriter_optimized.xlsx']

for file in file_list:
    try:
        wb = load_workbook(file)
        ws = wb.active

        value = ws['A1'].value
        print(f"{file}: A1 = {value}")
    except Exception as e:
        print("Bład", e)

# .\dane_zaktualizowane.xlsx: A1 = Rank
# .\openpyxl_optimized.xlsx: A1 = 0
# .\tabela_przetawna.xlsx: A1 = Data
# .\vgsales_formated.xlsx: A1 = Rank
# .\videogamesales.xlsx: A1 = Rank
# .\xlswriter_optimized.xlsx: A1 = 0
