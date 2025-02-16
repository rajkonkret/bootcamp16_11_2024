from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook("videogamesales.xlsx")

ws = wb["vgsales"]

# pogrubienie pierwszego wiersza
# for cell in ws[1]:
#     cell.font = Font(bold=True)

for cell in ws["1:1"]:
    cell.font = Font(bold=True, size=12)

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter

    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            print("Bład")

        ws.column_dimensions[col_letter].width = max_length + 2
wb.save("vgsales_formated.xlsx")